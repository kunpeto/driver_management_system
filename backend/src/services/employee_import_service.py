"""
EmployeeImportService 員工批次匯入服務
對應 tasks.md T048: 實作批次匯入服務

提供 Excel 批次匯入員工資料功能。
"""

import io
from dataclasses import dataclass, field
from typing import BinaryIO, Optional

from sqlalchemy.orm import Session

from src.services.employee_service import (
    DuplicateEmployeeError,
    EmployeeService,
    InvalidEmployeeIdError,
)


@dataclass
class ImportResult:
    """匯入結果"""
    success: bool
    total_rows: int = 0
    imported_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    errors: list[dict] = field(default_factory=list)
    imported_ids: list[str] = field(default_factory=list)


@dataclass
class RowValidationResult:
    """行驗證結果"""
    valid: bool
    row_number: int
    data: Optional[dict] = None
    error: Optional[str] = None


class EmployeeImportService:
    """
    員工批次匯入服務

    支援從 Excel 檔案批次匯入員工資料。

    Excel 格式要求：
    - 第一行為標題行
    - 必要欄位：員工編號、姓名、部門
    - 選填欄位：電話、電子郵件、緊急聯絡人、緊急聯絡電話
    """

    # 欄位名稱對照（支援多種名稱）
    COLUMN_MAPPING = {
        "employee_id": ["員工編號", "編號", "employee_id", "id"],
        "employee_name": ["姓名", "員工姓名", "name", "employee_name"],
        "current_department": ["部門", "現職部門", "department", "current_department"],
        "phone": ["電話", "手機", "phone", "mobile"],
        "email": ["電子郵件", "email", "e-mail", "郵件"],
        "emergency_contact": ["緊急聯絡人", "emergency_contact"],
        "emergency_phone": ["緊急聯絡電話", "emergency_phone"],
    }

    # 必要欄位
    REQUIRED_COLUMNS = ["employee_id", "employee_name", "current_department"]

    def __init__(self, db: Session):
        """
        初始化服務

        Args:
            db: SQLAlchemy Session
        """
        self.db = db
        self._employee_service = EmployeeService(db)

    def import_from_excel(
        self,
        file: BinaryIO,
        skip_duplicates: bool = True,
        created_by: str = "import"
    ) -> ImportResult:
        """
        從 Excel 檔案匯入員工

        Args:
            file: Excel 檔案（二進位串流）
            skip_duplicates: 是否跳過重複的員工編號
            created_by: 操作者名稱

        Returns:
            ImportResult: 匯入結果
        """
        try:
            import openpyxl
        except ImportError:
            return ImportResult(
                success=False,
                errors=[{"row": 0, "error": "需要安裝 openpyxl 套件才能匯入 Excel"}]
            )

        result = ImportResult(success=True)

        try:
            # 讀取 Excel
            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet = workbook.active

            if sheet is None:
                return ImportResult(
                    success=False,
                    errors=[{"row": 0, "error": "Excel 檔案沒有工作表"}]
                )

            rows = list(sheet.iter_rows(values_only=True))

            if len(rows) < 2:
                return ImportResult(
                    success=False,
                    errors=[{"row": 0, "error": "Excel 檔案至少需要標題行和一筆資料"}]
                )

            # 解析標題行
            header_row = rows[0]
            column_indices = self._parse_header(header_row)

            # 檢查必要欄位
            missing_columns = [
                col for col in self.REQUIRED_COLUMNS
                if col not in column_indices
            ]
            if missing_columns:
                return ImportResult(
                    success=False,
                    errors=[{
                        "row": 1,
                        "error": f"缺少必要欄位：{', '.join(missing_columns)}"
                    }]
                )

            result.total_rows = len(rows) - 1  # 不含標題行

            # 處理資料行（使用批次模式提升效能）
            for row_idx, row in enumerate(rows[1:], start=2):
                validation = self._validate_row(row, column_indices, row_idx)

                if not validation.valid:
                    result.error_count += 1
                    result.errors.append({
                        "row": row_idx,
                        "error": validation.error
                    })
                    continue

                # 嘗試建立員工（auto_commit=False 避免每筆都 commit）
                try:
                    employee = self._employee_service.create(
                        employee_id=validation.data["employee_id"],
                        employee_name=validation.data["employee_name"],
                        current_department=validation.data["current_department"],
                        phone=validation.data.get("phone"),
                        email=validation.data.get("email"),
                        emergency_contact=validation.data.get("emergency_contact"),
                        emergency_phone=validation.data.get("emergency_phone"),
                        auto_commit=False  # 批次模式：延遲 commit
                    )
                    result.imported_count += 1
                    result.imported_ids.append(employee.employee_id)

                except DuplicateEmployeeError:
                    if skip_duplicates:
                        result.skipped_count += 1
                    else:
                        result.error_count += 1
                        result.errors.append({
                            "row": row_idx,
                            "error": f"員工編號已存在：{validation.data['employee_id']}"
                        })

                except InvalidEmployeeIdError as e:
                    result.error_count += 1
                    result.errors.append({
                        "row": row_idx,
                        "error": str(e)
                    })

                except Exception as e:
                    result.error_count += 1
                    result.errors.append({
                        "row": row_idx,
                        "error": f"建立員工失敗：{str(e)}"
                    })

            # 批次提交所有變更（效能優化：單次 commit）
            if result.imported_count > 0:
                self._employee_service.commit()

            # 設定整體結果
            result.success = result.error_count == 0

        except Exception as e:
            return ImportResult(
                success=False,
                errors=[{"row": 0, "error": f"讀取 Excel 失敗：{str(e)}"}]
            )

        return result

    def validate_excel(self, file: BinaryIO) -> ImportResult:
        """
        驗證 Excel 檔案格式（不實際匯入）

        Args:
            file: Excel 檔案

        Returns:
            ImportResult: 驗證結果
        """
        try:
            import openpyxl
        except ImportError:
            return ImportResult(
                success=False,
                errors=[{"row": 0, "error": "需要安裝 openpyxl 套件"}]
            )

        result = ImportResult(success=True)

        try:
            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet = workbook.active

            if sheet is None:
                return ImportResult(
                    success=False,
                    errors=[{"row": 0, "error": "Excel 檔案沒有工作表"}]
                )

            rows = list(sheet.iter_rows(values_only=True))

            if len(rows) < 2:
                return ImportResult(
                    success=False,
                    errors=[{"row": 0, "error": "Excel 檔案至少需要標題行和一筆資料"}]
                )

            # 解析標題行
            header_row = rows[0]
            column_indices = self._parse_header(header_row)

            # 檢查必要欄位
            missing_columns = [
                col for col in self.REQUIRED_COLUMNS
                if col not in column_indices
            ]
            if missing_columns:
                return ImportResult(
                    success=False,
                    errors=[{
                        "row": 1,
                        "error": f"缺少必要欄位：{', '.join(missing_columns)}"
                    }]
                )

            result.total_rows = len(rows) - 1

            # 驗證每一行
            for row_idx, row in enumerate(rows[1:], start=2):
                validation = self._validate_row(row, column_indices, row_idx)
                if not validation.valid:
                    result.error_count += 1
                    result.errors.append({
                        "row": row_idx,
                        "error": validation.error
                    })

            result.success = result.error_count == 0

        except Exception as e:
            return ImportResult(
                success=False,
                errors=[{"row": 0, "error": f"讀取 Excel 失敗：{str(e)}"}]
            )

        return result

    def get_template_columns(self) -> list[dict]:
        """
        取得匯入範本欄位資訊

        Returns:
            list[dict]: 欄位資訊列表
        """
        return [
            {
                "name": "員工編號",
                "key": "employee_id",
                "required": True,
                "description": "員工編號（如 1011M0095）"
            },
            {
                "name": "姓名",
                "key": "employee_name",
                "required": True,
                "description": "員工姓名"
            },
            {
                "name": "部門",
                "key": "current_department",
                "required": True,
                "description": "部門（淡海 或 安坑）"
            },
            {
                "name": "電話",
                "key": "phone",
                "required": False,
                "description": "聯絡電話"
            },
            {
                "name": "電子郵件",
                "key": "email",
                "required": False,
                "description": "電子郵件"
            },
            {
                "name": "緊急聯絡人",
                "key": "emergency_contact",
                "required": False,
                "description": "緊急聯絡人姓名"
            },
            {
                "name": "緊急聯絡電話",
                "key": "emergency_phone",
                "required": False,
                "description": "緊急聯絡電話"
            }
        ]

    def _parse_header(self, header_row: tuple) -> dict[str, int]:
        """
        解析標題行，建立欄位名稱到索引的對照

        Args:
            header_row: 標題行資料

        Returns:
            dict: 欄位名稱 -> 欄位索引
        """
        column_indices = {}

        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue

            cell_value = str(cell_value).strip().lower()

            for field_name, aliases in self.COLUMN_MAPPING.items():
                if cell_value in [a.lower() for a in aliases]:
                    column_indices[field_name] = idx
                    break

        return column_indices

    def _validate_row(
        self,
        row: tuple,
        column_indices: dict[str, int],
        row_number: int
    ) -> RowValidationResult:
        """
        驗證資料行

        Args:
            row: 資料行
            column_indices: 欄位索引對照
            row_number: 行號

        Returns:
            RowValidationResult: 驗證結果
        """
        data = {}

        # 提取資料
        for field_name, col_idx in column_indices.items():
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None:
                    data[field_name] = str(value).strip()

        # 檢查必要欄位
        for required in self.REQUIRED_COLUMNS:
            if required not in data or not data[required]:
                return RowValidationResult(
                    valid=False,
                    row_number=row_number,
                    error=f"缺少必要欄位：{required}"
                )

        # 驗證部門
        valid_departments = ["淡海", "安坑"]
        if data["current_department"] not in valid_departments:
            return RowValidationResult(
                valid=False,
                row_number=row_number,
                error=f"無效的部門：{data['current_department']}，應為 淡海 或 安坑"
            )

        return RowValidationResult(
            valid=True,
            row_number=row_number,
            data=data
        )
