"""
EmployeeExportService 員工批次匯出服務
對應 tasks.md T049: 實作批次匯出服務

提供 Excel 批次匯出員工資料功能。
"""

import io
from typing import BinaryIO, Optional

from sqlalchemy.orm import Session

from src.models.employee import Employee
from src.models.google_oauth_token import Department
from src.services.employee_service import EmployeeService


class EmployeeExportService:
    """
    員工批次匯出服務

    支援將員工資料匯出為 Excel 檔案。
    """

    # 匯出欄位定義
    EXPORT_COLUMNS = [
        {"key": "employee_id", "header": "員工編號", "width": 15},
        {"key": "employee_name", "header": "姓名", "width": 12},
        {"key": "current_department", "header": "部門", "width": 10},
        {"key": "hire_year_month", "header": "入職年月", "width": 12},
        {"key": "phone", "header": "電話", "width": 15},
        {"key": "email", "header": "電子郵件", "width": 25},
        {"key": "emergency_contact", "header": "緊急聯絡人", "width": 12},
        {"key": "emergency_phone", "header": "緊急聯絡電話", "width": 15},
        {"key": "is_resigned", "header": "離職狀態", "width": 10},
    ]

    def __init__(self, db: Session):
        """
        初始化服務

        Args:
            db: SQLAlchemy Session
        """
        self.db = db
        self._employee_service = EmployeeService(db)

    def export_to_excel(
        self,
        department: Optional[str] = None,
        include_resigned: bool = False,
        search: Optional[str] = None
    ) -> BinaryIO:
        """
        匯出員工資料為 Excel

        Args:
            department: 篩選部門（選填）
            include_resigned: 是否包含離職員工
            search: 搜尋關鍵字（選填）

        Returns:
            BinaryIO: Excel 檔案二進位串流
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("需要安裝 openpyxl 套件才能匯出 Excel")

        # 取得員工資料
        employees = self._employee_service.list_all(
            department=department,
            include_resigned=include_resigned,
            search=search,
            limit=10000
        )

        # 建立工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "員工資料"

        # 設定標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 寫入標題行
        for col_idx, col_def in enumerate(self.EXPORT_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=col_def["header"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # 設定欄位寬度
            sheet.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

        # 寫入資料行
        cell_alignment = Alignment(vertical="center")

        for row_idx, employee in enumerate(employees, start=2):
            for col_idx, col_def in enumerate(self.EXPORT_COLUMNS, start=1):
                value = getattr(employee, col_def["key"], "")

                # 特殊處理
                if col_def["key"] == "is_resigned":
                    value = "離職" if value else "在職"

                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 凍結標題行
        sheet.freeze_panes = "A2"

        # 儲存到記憶體
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    def export_to_csv(
        self,
        department: Optional[str] = None,
        include_resigned: bool = False,
        search: Optional[str] = None
    ) -> str:
        """
        匯出員工資料為 CSV

        Args:
            department: 篩選部門（選填）
            include_resigned: 是否包含離職員工
            search: 搜尋關鍵字（選填）

        Returns:
            str: CSV 內容
        """
        import csv
        import io

        # 取得員工資料
        employees = self._employee_service.list_all(
            department=department,
            include_resigned=include_resigned,
            search=search,
            limit=10000
        )

        output = io.StringIO()
        writer = csv.writer(output)

        # 寫入標題行
        headers = [col["header"] for col in self.EXPORT_COLUMNS]
        writer.writerow(headers)

        # 寫入資料行
        for employee in employees:
            row = []
            for col_def in self.EXPORT_COLUMNS:
                value = getattr(employee, col_def["key"], "")

                # 特殊處理
                if col_def["key"] == "is_resigned":
                    value = "離職" if value else "在職"

                row.append(value or "")

            writer.writerow(row)

        return output.getvalue()

    def export_template(self) -> BinaryIO:
        """
        匯出匯入範本

        Returns:
            BinaryIO: Excel 範本檔案
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError:
            raise RuntimeError("需要安裝 openpyxl 套件")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "員工匯入範本"

        # 標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        required_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 範本欄位定義
        template_columns = [
            {"header": "員工編號", "width": 15, "required": True, "example": "1011M0095"},
            {"header": "姓名", "width": 12, "required": True, "example": "張三"},
            {"header": "部門", "width": 10, "required": True, "example": "淡海"},
            {"header": "電話", "width": 15, "required": False, "example": "0912345678"},
            {"header": "電子郵件", "width": 25, "required": False, "example": "example@mail.com"},
            {"header": "緊急聯絡人", "width": 12, "required": False, "example": "李四"},
            {"header": "緊急聯絡電話", "width": 15, "required": False, "example": "0987654321"},
        ]

        # 寫入標題行
        for col_idx, col_def in enumerate(template_columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=col_def["header"])
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

            if col_def["required"]:
                cell.fill = required_fill
            else:
                cell.fill = header_fill

            sheet.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

        # 寫入範例資料
        for col_idx, col_def in enumerate(template_columns, start=1):
            cell = sheet.cell(row=2, column=col_idx, value=col_def["example"])
            cell.border = thin_border

        # 部門下拉選單（第 3 欄）
        dept_validation = DataValidation(
            type="list",
            formula1='"淡海,安坑"',
            allow_blank=False
        )
        dept_validation.error = "請選擇有效的部門"
        dept_validation.errorTitle = "無效的部門"
        sheet.add_data_validation(dept_validation)
        dept_validation.add(sheet["C2:C1000"])

        # 凍結標題行
        sheet.freeze_panes = "A2"

        # 加入說明工作表
        help_sheet = workbook.create_sheet(title="說明")
        help_content = [
            ["員工匯入範本說明"],
            [""],
            ["必填欄位（黃色標題）："],
            ["  - 員工編號：格式為 YYMM + 類型碼 + 4位序號（如 1011M0095）"],
            ["  - 姓名：員工姓名"],
            ["  - 部門：淡海 或 安坑"],
            [""],
            ["選填欄位（藍色標題）："],
            ["  - 電話：聯絡電話"],
            ["  - 電子郵件：電子郵件地址"],
            ["  - 緊急聯絡人：緊急聯絡人姓名"],
            ["  - 緊急聯絡電話：緊急聯絡電話"],
            [""],
            ["注意事項："],
            ["  1. 第一行為標題行，請勿修改"],
            ["  2. 員工編號不可重複"],
            ["  3. 部門只能填寫「淡海」或「安坑」"],
            ["  4. 入職年月會自動從員工編號解析"],
        ]

        for row_idx, content in enumerate(help_content, start=1):
            if isinstance(content, list):
                sheet_cell = help_sheet.cell(row=row_idx, column=1, value=content[0] if content else "")
            else:
                sheet_cell = help_sheet.cell(row=row_idx, column=1, value=content)

        help_sheet.column_dimensions["A"].width = 60

        # 儲存到記憶體
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    def get_export_count(
        self,
        department: Optional[str] = None,
        include_resigned: bool = False,
        search: Optional[str] = None
    ) -> int:
        """
        取得匯出資料筆數

        Args:
            department: 篩選部門
            include_resigned: 是否包含離職員工
            search: 搜尋關鍵字

        Returns:
            int: 資料筆數
        """
        return self._employee_service.count(
            department=department,
            include_resigned=include_resigned,
            search=search
        )
